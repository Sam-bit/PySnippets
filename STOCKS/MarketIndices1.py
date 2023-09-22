import json
import os.path
import time
import urllib.parse
from openpyxl.styles import Font
import openpyxl
import requests
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
headers = {
    'Connection': 'keep-alive',
    'Cache-Control': 'max-age=0',
    'DNT': '1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36',
    'Sec-Fetch-User': '?1',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-Mode': 'navigate',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'en-US,en;q=0.9,hi;q=0.8',
}
equitycookies = {
    '_ga': 'GA1.1.1907533281.1681102554',
    'defaultLang': 'en',
    'ext_name': 'ojplmecpdpgccookcobabopnaifgidhf',
    'nsit': 'iQLsuGDvx8sZkJDSSlgrkS4P',
    'nseappid': 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJhcGkubnNlIiwiYXVkIjoiYXBpLm5zZSIsImlhdCI6MTY4MTEyOTc3MywiZXhwIjoxNjgxMTM2OTczfQ.yjsWWUssGysYhDitpgeYETqs9_aCPWcawDkh0U1-FXw',
    'bm_mi': 'A63270490E2907A6BCB3171633363160~YAAQLbYRYDr8CmGHAQAARJwkaxN+6EFsPpSI+acEErk3oeICZ2yzc6krD95P0p+npLFE10AirjUw4MwPnf+mFBPXWCVPlITbrgECbGgldxuNnLmRuNGLkXS0g+BBMGPVtuzTFYJVSW8sVsmWRc740NTf7VxiW/OMWdFIZLmJOmjffqPkAxOJHV0Fj4qWmz0XHaYegB93lte+O3c4wtD84Bm/SofhaVsLxiwMczKpK8DCogHn4pMpE+muPXtQC90PVAAZVIh7YCobpXSYrRk0u2hqXuKWu9WtLfgNqtrkzDAt6D2h6EFDtrO4/Pc4mOyr5Fg71u7e2u5f1dYoEZjddO/iLZpEG4LjuEGk6w==~1',
    '_ga_PJSKY6CFJH': 'GS1.1.1681129772.4.1.1681129778.54.0.0',
    'ak_bmsc': '59B3CCF799604334611D20AB281D6333~000000000000000000000000000000~YAAQLbYRYAf/CmGHAQAAyLokaxM+XgPIxrTIPsSmvKEe2gXV1LaBV1I1nLAGyP+ZGxDQeMQmncpLwJTM5K/W68ifOlUHdj+il+bsPc5DKK1vSA+3/p4sfZAklO/8f6ArJOIrf1lySd5cYKd+kjsZkFdfBRmZaCQfbYSPhkthRrGR0Cn4ljysmzI6+LrDgOhgM5gJuK0gGWeTCawjvMAAiKfKJPXhSEl0A5Mxluzy5oXZEiB4MzWvU4RVaY2cl2/QwZMB/NMQUd1K6ihIS7PTKbizOClQEOzS1zLWbJds9qwVFvtvTOaHs6ASHKZYRb7bkXoEfHljQgH+GXT4NZ0e+trNLyRCMzh/htIQ5mgyl+RszJ0HH9HvSUQL7Ng6UXubdWPBiUKsBzVUjmSFNlXIqXUxMYdhn+TUbi4kucfk6gbPGQIcQcyezsn3apCR+J95rRgXCvcLFT6lAsmULnjg',
    'bm_sv': 'AC23330D1ECEA0CF30F2945510894B23~YAAQLbYRYAb/DWGHAQAAw15VaxO4QwyPoB/3WaaY6c+xStVOVxktOWXCiBTfO8pxpgdrzwNWfE2eIZaZWAHP0yqyZN8lorouf/Rin0lkOJX0Dav1k7ocd09HMP50RSev4WXb/q/VMg0yGubb7y3MDweiXk8ZlgroYGxVIGmn5/A/zx+xGVPJrXT1gSL+I6MAhgMwUGq9EiiHUKwr8OS9jme5iAhqjqfKDgdJfa+ieqBg+jRhiyc6BNu8vYB6z0DoaXO1~1',
    'RT': '"z=1&dm=nseindia.com&si=573ca877-6b92-4c5a-acd8-76ca8995445a&ss=lgagrcfk&sl=0&tt=0&bcn=%2F%2F684d0d4a.akstat.io%2F"',
}

equityheaders = {
    'authority': 'www.nseindia.com',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'accept-language': 'en-US,en;q=0.9',
    'cache-control': 'max-age=0',
    # 'cookie': '_ga=GA1.1.1907533281.1681102554; defaultLang=en; ext_name=ojplmecpdpgccookcobabopnaifgidhf; nsit=iQLsuGDvx8sZkJDSSlgrkS4P; nseappid=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJhcGkubnNlIiwiYXVkIjoiYXBpLm5zZSIsImlhdCI6MTY4MTEyOTc3MywiZXhwIjoxNjgxMTM2OTczfQ.yjsWWUssGysYhDitpgeYETqs9_aCPWcawDkh0U1-FXw; bm_mi=A63270490E2907A6BCB3171633363160~YAAQLbYRYDr8CmGHAQAARJwkaxN+6EFsPpSI+acEErk3oeICZ2yzc6krD95P0p+npLFE10AirjUw4MwPnf+mFBPXWCVPlITbrgECbGgldxuNnLmRuNGLkXS0g+BBMGPVtuzTFYJVSW8sVsmWRc740NTf7VxiW/OMWdFIZLmJOmjffqPkAxOJHV0Fj4qWmz0XHaYegB93lte+O3c4wtD84Bm/SofhaVsLxiwMczKpK8DCogHn4pMpE+muPXtQC90PVAAZVIh7YCobpXSYrRk0u2hqXuKWu9WtLfgNqtrkzDAt6D2h6EFDtrO4/Pc4mOyr5Fg71u7e2u5f1dYoEZjddO/iLZpEG4LjuEGk6w==~1; _ga_PJSKY6CFJH=GS1.1.1681129772.4.1.1681129778.54.0.0; ak_bmsc=59B3CCF799604334611D20AB281D6333~000000000000000000000000000000~YAAQLbYRYAf/CmGHAQAAyLokaxM+XgPIxrTIPsSmvKEe2gXV1LaBV1I1nLAGyP+ZGxDQeMQmncpLwJTM5K/W68ifOlUHdj+il+bsPc5DKK1vSA+3/p4sfZAklO/8f6ArJOIrf1lySd5cYKd+kjsZkFdfBRmZaCQfbYSPhkthRrGR0Cn4ljysmzI6+LrDgOhgM5gJuK0gGWeTCawjvMAAiKfKJPXhSEl0A5Mxluzy5oXZEiB4MzWvU4RVaY2cl2/QwZMB/NMQUd1K6ihIS7PTKbizOClQEOzS1zLWbJds9qwVFvtvTOaHs6ASHKZYRb7bkXoEfHljQgH+GXT4NZ0e+trNLyRCMzh/htIQ5mgyl+RszJ0HH9HvSUQL7Ng6UXubdWPBiUKsBzVUjmSFNlXIqXUxMYdhn+TUbi4kucfk6gbPGQIcQcyezsn3apCR+J95rRgXCvcLFT6lAsmULnjg; bm_sv=AC23330D1ECEA0CF30F2945510894B23~YAAQLbYRYAb/DWGHAQAAw15VaxO4QwyPoB/3WaaY6c+xStVOVxktOWXCiBTfO8pxpgdrzwNWfE2eIZaZWAHP0yqyZN8lorouf/Rin0lkOJX0Dav1k7ocd09HMP50RSev4WXb/q/VMg0yGubb7y3MDweiXk8ZlgroYGxVIGmn5/A/zx+xGVPJrXT1gSL+I6MAhgMwUGq9EiiHUKwr8OS9jme5iAhqjqfKDgdJfa+ieqBg+jRhiyc6BNu8vYB6z0DoaXO1~1; RT="z=1&dm=nseindia.com&si=573ca877-6b92-4c5a-acd8-76ca8995445a&ss=lgagrcfk&sl=0&tt=0&bcn=%2F%2F684d0d4a.akstat.io%2F"',
    'dnt': '1',
    'sec-ch-ua': '"Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'none',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36',
}

def openSeleniumUrl(siteurl):
    from selenium import webdriver
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    #options.add_argument("--headless")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(options=options, service=Service(r'C:\DRIVERS\chromedriver.exe'))
    driver.get(siteurl)
    delay = 9  # seconds
    try:
        myElem = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.ID, 'sectoralindices')))
        all_cookies = driver.get_cookies()
        cookies_dict = {}
        for cookie in all_cookies:
            cookies_dict[cookie['name']] = cookie['value']
        equitycookies = cookies_dict
        driver.close()
    except TimeoutException:
        print("Loading took too much time!")
def getSeleniumUrlContent(siteurl):
    from selenium import webdriver
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    #options.add_argument("--headless")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    driver = webdriver.Chrome(options=options, service=Service(r'C:\DRIVERS\chromedriver.exe'))
    driver.get(siteurl)
    WebDriverWait(driver,10)
    import re, html
    tag_re = re.compile(r'(<!--.*?-->|<[^>]*>)')
    no_tags = tag_re.sub('', driver.page_source)
    return no_tags
def Sort_List(sub_li,index,descending = True):
    sub_li.sort(key=lambda x: x[index],reverse = descending)
    return sub_li
r = requests.get('https://www.nseindia.com/json/live-index.json',headers = headers)
#print(r.content)
jsonText = r.content
resp = json.loads(jsonText)
indicescolumns = [str(resp['columns'][i]['heading']).upper() for i in range(len(resp['columns']))]
openSeleniumUrl('https://www.nseindia.com/market-data/live-market-indices')
r = getSeleniumUrlContent('https://www.nseindia.com/api/allIndices')
#print(r.content)
jsonText = r
print(jsonText)
resp = json.loads(jsonText)
#INDEX,CURRENT,%CHNG,OPEN,HIGH,LOW,PREV. CLOSE,PREV. DAY,1W AGO,1M AGO,1Y AGO,52W H,52W L,TODAY
print(resp['data'][0])
marketindices = [
    [
        resp['data'][i]['key'],
        resp['data'][i]['index'],
        resp['data'][i]['last'],
        resp['data'][i]['percentChange'],
        resp['data'][i]['open'],
        resp['data'][i]['high'],
        resp['data'][i]['low'],
        resp['data'][i]['previousClose'],
        resp['data'][i]['previousDay'],
        resp['data'][i].get('oneWeekAgo',0),
        resp['data'][i].get('oneMonthAgo',0),
        resp['data'][i].get('oneYearAgo',0),
        resp['data'][i]['yearHigh'],
        resp['data'][i]['yearLow']
        ] for i in range(len(resp['data']))
    ]

header_broad_market_indices = [
'','BROAD MARKET INDICES','','','','','','','','','','','','']
temp_broad_market_indices = Sort_List([
    [
        resp['data'][i]['key'],
        resp['data'][i]['index'],
        resp['data'][i]['last'],
        resp['data'][i]['percentChange'],
        resp['data'][i]['open'],
        resp['data'][i]['high'],
        resp['data'][i]['low'],
        resp['data'][i]['previousClose'],
        resp['data'][i]['previousDay'],
        resp['data'][i].get('oneWeekAgo', 0),
        resp['data'][i].get('oneMonthAgo', 0),
        resp['data'][i].get('oneYearAgo', 0),
        resp['data'][i]['yearHigh'],
        resp['data'][i]['yearLow']
        ] for i in range(len(resp['data'])) if resp['data'][i]['key'] == 'BROAD MARKET INDICES'
    ],3,True)
temp_broad_market_indices.insert(0,header_broad_market_indices)

header_sectoral_market_indices = [
'','SECTORAL INDICES','','','','','','','','','','','','']
temp_sectoral_market_indices = Sort_List([
    [
        resp['data'][i]['key'],
        resp['data'][i]['index'],
        resp['data'][i]['last'],
        resp['data'][i]['percentChange'],
        resp['data'][i]['open'],
        resp['data'][i]['high'],
        resp['data'][i]['low'],
        resp['data'][i]['previousClose'],
        resp['data'][i]['previousDay'],
        resp['data'][i].get('oneWeekAgo', 0),
        resp['data'][i].get('oneMonthAgo', 0),
        resp['data'][i].get('oneYearAgo', 0),
        resp['data'][i]['yearHigh'],
        resp['data'][i]['yearLow']
        ] for i in range(len(resp['data'])) if resp['data'][i]['key'] == 'SECTORAL INDICES'
    ],3,True)
temp_sectoral_market_indices.insert(0,header_sectoral_market_indices)

header_strategy_market_indices = [
'','STRATEGY INDICES','','','','','','','','','','','','']
temp_strategy_market_indices = Sort_List([
    [
        resp['data'][i]['key'],
        resp['data'][i]['index'],
        resp['data'][i]['last'],
        resp['data'][i]['percentChange'],
        resp['data'][i]['open'],
        resp['data'][i]['high'],
        resp['data'][i]['low'],
        resp['data'][i]['previousClose'],
        resp['data'][i]['previousDay'],
        resp['data'][i].get('oneWeekAgo', 0),
        resp['data'][i].get('oneMonthAgo', 0),
        resp['data'][i].get('oneYearAgo', 0),
        resp['data'][i]['yearHigh'],
        resp['data'][i]['yearLow']
        ] for i in range(len(resp['data'])) if resp['data'][i]['key'] == 'STRATEGY INDICES'
    ],3,True)
temp_strategy_market_indices.insert(0,header_strategy_market_indices)

header_thematic_market_indices = [
'','THEMATIC INDICES','','','','','','','','','','','','']
temp_thematic_market_indices = Sort_List([
    [
        resp['data'][i]['key'],
        resp['data'][i]['index'],
        resp['data'][i]['last'],
        resp['data'][i]['percentChange'],
        resp['data'][i]['open'],
        resp['data'][i]['high'],
        resp['data'][i]['low'],
        resp['data'][i]['previousClose'],
        resp['data'][i]['previousDay'],
        resp['data'][i].get('oneWeekAgo', 0),
        resp['data'][i].get('oneMonthAgo', 0),
        resp['data'][i].get('oneYearAgo', 0),
        resp['data'][i]['yearHigh'],
        resp['data'][i]['yearLow']
        ] for i in range(len(resp['data'])) if resp['data'][i]['key'] == 'THEMATIC INDICES'
    ],3,True)
temp_thematic_market_indices.insert(0,header_thematic_market_indices)

header_fixedincome_market_indices = [
'','FIXED INCOME INDICES','','','','','','','','','','','','']
temp_fixedincome_market_indices = Sort_List([
    [
        resp['data'][i]['key'],
        resp['data'][i]['index'],
        resp['data'][i]['last'],
        resp['data'][i]['percentChange'],
        resp['data'][i]['open'],
        resp['data'][i]['high'],
        resp['data'][i]['low'],
        resp['data'][i]['previousClose'],
        resp['data'][i]['previousDay'],
        resp['data'][i].get('oneWeekAgo', 0),
        resp['data'][i].get('oneMonthAgo', 0),
        resp['data'][i].get('oneYearAgo', 0),
        resp['data'][i]['yearHigh'],
        resp['data'][i]['yearLow']
        ] for i in range(len(resp['data'])) if resp['data'][i]['key'] == 'FIXED INCOME INDICES'
    ],3,True)
temp_fixedincome_market_indices.insert(0,header_fixedincome_market_indices)

new_market_indices =[]
new_market_indices.extend(temp_broad_market_indices)
new_market_indices.extend(temp_sectoral_market_indices)
new_market_indices.extend(temp_strategy_market_indices)
new_market_indices.extend(temp_thematic_market_indices)
new_market_indices.extend(temp_fixedincome_market_indices)
print(new_market_indices)
import datetime
current_datetime = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
xlfilename = "C:\\Users\\LENOVO\\Desktop\\MARKET_INDICES\\MARKET_INDICES_"+current_datetime+".xlsx"
my_wb= openpyxl.Workbook()
my_sheet_obj = my_wb.active
my_sheet_obj.title = 'INDICES'
for i in range(len(new_market_indices)):
    for j in range(len(new_market_indices[0])):
        if new_market_indices[i][0]!= "":
            my_sheet_obj.cell(row=i+1,column=2).hyperlink = '#'+new_market_indices[i][1]+'!A1'
            my_sheet_obj.cell(row=i+1,column=2).font = Font(underline='single', color='0563C1')
        else:
            my_sheet_obj.cell(row=i + 1, column=2).font = Font(bold=True)
        my_sheet_obj.cell(row=i+1,column=j+1).value = new_market_indices[i][j]
my_wb.save(xlfilename)

r = requests.get('https://www.nseindia.com/json/equity-stockIndices.json', headers=headers)
# print(r.content)
jsonText = r.content
resp = json.loads(jsonText)
equitycolumns = [str(resp['columns'][i]['heading']).upper() for i in range(len(resp['columns']))]
for i in range(len(marketindices)):
    if marketindices[i][0] != 'FIXED INCOME INDICES' and marketindices[i][1] not in('NIFTY100 ESG SECTOR LEADERS','NIFTY50 TR 2X LEVERAGE','NIFTY50 PR 2X LEVERAGE','NIFTY50 TR 1X INVERSE','NIFTY50 PR 1X INVERSE','NIFTY50 DIVIDEND POINTS','INDIA VIX'):
        #print(marketindices[i])
        equityurl = 'https://www.nseindia.com/api/equity-stockIndices?index='+urllib.parse.quote(marketindices[i][1])
        print(equityurl)
        jsonText = getSeleniumUrlContent(equityurl)
        resp = json.loads(jsonText)
        headerindices = [resp['name'] + "---" + resp['timestamp']]
        headerindicesdata = resp['data'][0]
        equityindices = Sort_List([
            [
                resp['data'][i]['symbol'],
                resp['data'][i]['open'],
                resp['data'][i]['dayHigh'],
                resp['data'][i]['dayLow'],
                resp['data'][i]['previousClose'],
                resp['data'][i]['lastPrice'],
                resp['data'][i]['change'],
                resp['data'][i]['pChange'],
                resp['data'][i]['totalTradedVolume'],
                resp['data'][i]['totalTradedValue'],
                resp['data'][i]['yearHigh'],
                resp['data'][i]['yearLow']
            ] for i in range(1, len(resp['data']))
        ], 7, True)
        new_market_indices = []
        new_market_indices.extend(headerindices)
        new_market_indices.extend(headerindicesdata)
        new_market_indices.extend(equityindices)
        print(marketindices[i][1])
        my_wb.create_sheet(marketindices[i][1])
        for i in range(len(new_market_indices)):
            for j in range(len(new_market_indices[0])):
                my_sheet_obj.cell(row=i + 1, column=j + 1).value = new_market_indices[i][j]
        my_wb.save(xlfilename)
        '''
        r = requests.get(equityurl, cookies=equitycookies, headers=equityheaders)
        print(r.content)
        r.raise_for_status()  # raises exception when not a 2xx response
        if r.status_code != 204:
            jsonText = r.content
            resp = json.loads(jsonText)
            headerindices = [resp['name']+"---"+resp['timestamp']]
            headerindicesdata=resp['data'][0]
            equityindices = Sort_List([
            [
                resp['data'][i]['symbol'],
                resp['data'][i]['open'],
                resp['data'][i]['dayHigh'],
                resp['data'][i]['dayLow'],
                resp['data'][i]['previousClose'],
                resp['data'][i]['lastPrice'],
                resp['data'][i]['change'],
                resp['data'][i]['pChange'],
                resp['data'][i]['totalTradedVolume'],
                resp['data'][i]['totalTradedValue'],
                resp['data'][i]['yearHigh'],
                resp['data'][i]['yearLow']
            ] for i in range(1,len(resp['data']))
            ],7,True)
            new_market_indices = []
            new_market_indices.extend(headerindices)
            new_market_indices.extend(headerindicesdata)
            new_market_indices.extend(equityindices)
            print(marketindices[i][1])
            my_wb.create_sheet(marketindices[i][1])
            for i in range(len(new_market_indices)):
                for j in range(len(new_market_indices[0])):
                    my_sheet_obj.cell(row=i + 1, column=j + 1).value = new_market_indices[i][j]
            my_wb.save(xlfilename)
        '''
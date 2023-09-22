import json
import urllib

import openpyxl
import requests
from openpyxl.styles import Font

from STOCKS.PyNSE import nsefetch, headers
def Sort_List(sub_li,index,descending = True):
    sub_li.sort(key=lambda x: x[index],reverse = descending)
    return sub_li
r = requests.get('https://www.nseindia.com/json/live-index.json',headers = headers)
#print(r.content)
jsonText = r.content
payload = json.loads(jsonText)
indicescolumns = [str(payload['columns'][i]['heading']).upper() for i in range(len(payload['columns']))]
payload = nsefetch("https://www.nseindia.com/api/allIndices")
marketindices = [
    [
        payload['data'][i]['key'],
        payload['data'][i]['index'],
        payload['data'][i]['last'],
        payload['data'][i]['percentChange'],
        payload['data'][i]['open'],
        payload['data'][i]['high'],
        payload['data'][i]['low'],
        payload['data'][i]['previousClose'],
        payload['data'][i]['previousDay'],
        payload['data'][i].get('oneWeekAgo',0),
        payload['data'][i].get('oneMonthAgo',0),
        payload['data'][i].get('oneYearAgo',0),
        payload['data'][i]['yearHigh'],
        payload['data'][i]['yearLow']
        ] for i in range(len(payload['data']))
    ]

header_broad_market_indices = [
'','BROAD MARKET INDICES','','','','','','','','','','','','']
temp_broad_market_indices = Sort_List([
    [
        payload['data'][i]['key'],
        payload['data'][i]['index'],
        payload['data'][i]['last'],
        payload['data'][i]['percentChange'],
        payload['data'][i]['open'],
        payload['data'][i]['high'],
        payload['data'][i]['low'],
        payload['data'][i]['previousClose'],
        payload['data'][i]['previousDay'],
        payload['data'][i].get('oneWeekAgo', 0),
        payload['data'][i].get('oneMonthAgo', 0),
        payload['data'][i].get('oneYearAgo', 0),
        payload['data'][i]['yearHigh'],
        payload['data'][i]['yearLow']
        ] for i in range(len(payload['data'])) if payload['data'][i]['key'] == 'BROAD MARKET INDICES'
    ],3,True)
temp_broad_market_indices.insert(0,header_broad_market_indices)

header_sectoral_market_indices = [
'','SECTORAL INDICES','','','','','','','','','','','','']
temp_sectoral_market_indices = Sort_List([
    [
        payload['data'][i]['key'],
        payload['data'][i]['index'],
        payload['data'][i]['last'],
        payload['data'][i]['percentChange'],
        payload['data'][i]['open'],
        payload['data'][i]['high'],
        payload['data'][i]['low'],
        payload['data'][i]['previousClose'],
        payload['data'][i]['previousDay'],
        payload['data'][i].get('oneWeekAgo', 0),
        payload['data'][i].get('oneMonthAgo', 0),
        payload['data'][i].get('oneYearAgo', 0),
        payload['data'][i]['yearHigh'],
        payload['data'][i]['yearLow']
        ] for i in range(len(payload['data'])) if payload['data'][i]['key'] == 'SECTORAL INDICES'
    ],3,True)
temp_sectoral_market_indices.insert(0,header_sectoral_market_indices)

header_strategy_market_indices = [
'','STRATEGY INDICES','','','','','','','','','','','','']
temp_strategy_market_indices = Sort_List([
    [
        payload['data'][i]['key'],
        payload['data'][i]['index'],
        payload['data'][i]['last'],
        payload['data'][i]['percentChange'],
        payload['data'][i]['open'],
        payload['data'][i]['high'],
        payload['data'][i]['low'],
        payload['data'][i]['previousClose'],
        payload['data'][i]['previousDay'],
        payload['data'][i].get('oneWeekAgo', 0),
        payload['data'][i].get('oneMonthAgo', 0),
        payload['data'][i].get('oneYearAgo', 0),
        payload['data'][i]['yearHigh'],
        payload['data'][i]['yearLow']
        ] for i in range(len(payload['data'])) if payload['data'][i]['key'] == 'STRATEGY INDICES'
    ],3,True)
temp_strategy_market_indices.insert(0,header_strategy_market_indices)

header_thematic_market_indices = [
'','THEMATIC INDICES','','','','','','','','','','','','']
temp_thematic_market_indices = Sort_List([
    [
        payload['data'][i]['key'],
        payload['data'][i]['index'],
        payload['data'][i]['last'],
        payload['data'][i]['percentChange'],
        payload['data'][i]['open'],
        payload['data'][i]['high'],
        payload['data'][i]['low'],
        payload['data'][i]['previousClose'],
        payload['data'][i]['previousDay'],
        payload['data'][i].get('oneWeekAgo', 0),
        payload['data'][i].get('oneMonthAgo', 0),
        payload['data'][i].get('oneYearAgo', 0),
        payload['data'][i]['yearHigh'],
        payload['data'][i]['yearLow']
        ] for i in range(len(payload['data'])) if payload['data'][i]['key'] == 'THEMATIC INDICES'
    ],3,True)
temp_thematic_market_indices.insert(0,header_thematic_market_indices)

header_fixedincome_market_indices = [
'','FIXED INCOME INDICES','','','','','','','','','','','','']
temp_fixedincome_market_indices = Sort_List([
    [
        payload['data'][i]['key'],
        payload['data'][i]['index'],
        payload['data'][i]['last'],
        payload['data'][i]['percentChange'],
        payload['data'][i]['open'],
        payload['data'][i]['high'],
        payload['data'][i]['low'],
        payload['data'][i]['previousClose'],
        payload['data'][i]['previousDay'],
        payload['data'][i].get('oneWeekAgo', 0),
        payload['data'][i].get('oneMonthAgo', 0),
        payload['data'][i].get('oneYearAgo', 0),
        payload['data'][i]['yearHigh'],
        payload['data'][i]['yearLow']
        ] for i in range(len(payload['data'])) if payload['data'][i]['key'] == 'FIXED INCOME INDICES'
    ],3,True)
temp_fixedincome_market_indices.insert(0,header_fixedincome_market_indices)

new_market_indices =[]
new_market_indices.extend(temp_broad_market_indices)
new_market_indices.extend(temp_sectoral_market_indices)
new_market_indices.extend(temp_strategy_market_indices)
new_market_indices.extend(temp_thematic_market_indices)
new_market_indices.extend(temp_fixedincome_market_indices)
print(new_market_indices)
import datetime
current_datetime = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
xlfilename = "C:\\Users\\LENOVO\\Desktop\\MARKET_INDICES\\MARKET_INDICES_"+current_datetime+".xlsx"
my_wb= openpyxl.Workbook()
my_sheet_obj = my_wb.active
my_sheet_obj.title = 'INDICES'
for i in range(len(new_market_indices)):
    for j in range(len(new_market_indices[0])):
        if new_market_indices[i][0]!= "":
            my_sheet_obj.cell(row=i+1,column=2).hyperlink = '#'+new_market_indices[i][1]+'!A1'
            my_sheet_obj.cell(row=i+1,column=2).font = Font(underline='single', color='0563C1')
        else:
            my_sheet_obj.cell(row=i + 1, column=2).font = Font(bold=True)
        my_sheet_obj.cell(row=i+1,column=j+1).value = new_market_indices[i][j]
my_wb.save(xlfilename)

r = requests.get('https://www.nseindia.com/json/equity-stockIndices.json', headers=headers)
# print(r.content)
jsonText = r.content
resp = json.loads(jsonText)
equitycolumns = [str(resp['columns'][i]['heading']).upper() for i in range(len(resp['columns']))]
for i in range(len(marketindices)):
    if marketindices[i][0] != 'FIXED INCOME INDICES' and marketindices[i][1] not in('NIFTY100 ESG SECTOR LEADERS','NIFTY50 TR 2X LEVERAGE','NIFTY50 PR 2X LEVERAGE','NIFTY50 TR 1X INVERSE','NIFTY50 PR 1X INVERSE','NIFTY50 DIVIDEND POINTS','INDIA VIX'):
        #print(marketindices[i])
        equityurl = 'https://www.nseindia.com/api/equity-stockIndices?index='+urllib.parse.quote(marketindices[i][1])
        print(equityurl)
        position = nsefetch(equityurl)
        headerindices = [position['name'] + "---" + position['timestamp']]
        headerindicesdata = position['data'][0]
        equityindices = Sort_List([
            [
                position['data'][i]['symbol'],
                position['data'][i]['open'],
                position['data'][i]['dayHigh'],
                position['data'][i]['dayLow'],
                position['data'][i]['previousClose'],
                position['data'][i]['lastPrice'],
                position['data'][i]['change'],
                position['data'][i]['pChange'],
                position['data'][i]['totalTradedVolume'],
                position['data'][i]['totalTradedValue'],
                position['data'][i]['yearHigh'],
                position['data'][i]['yearLow']
            ] for i in range(1, len(position['data']))
        ], 7, True)
        new_market_indices = []
        new_market_indices.extend(equitycolumns)
        new_market_indices.extend(headerindices)
        new_market_indices.extend(headerindicesdata)
        new_market_indices.extend(equityindices)
        print(new_market_indices)
        my_wb.create_sheet(marketindices[i][1])
        idx_sheet_obj = my_wb[marketindices[i][1]]
        for i in range(len(new_market_indices)):
            for j in range(len(new_market_indices[0])):
                idx_sheet_obj.cell(row=i + 1, column=j + 1).value = new_market_indices[i][j]
        my_wb.save(xlfilename)
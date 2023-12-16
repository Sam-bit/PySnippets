import yfinance as yf
import openpyxl
import pandas_datareader as web
import requests
from bs4 import BeautifulSoup


def getSoup(symbol):
    headers = {
        'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'
    }
    url = "https://finance.yahoo.com/quote/" + symbol + "?p=" + symbol + "&.tsrc=fin-srch"
    html = requests.get(url, headers=headers).text
    soup = BeautifulSoup(html, 'html.parser')
    return soup


def get_ex_dividend_date(soup):
    try:
        exdivdate = soup.find('td', {"data-test": "EX_DIVIDEND_DATE-value"}).text.strip()
    except:
        exdivdate = "--"
    return exdivdate


def get_ex_dividend_amt_yield(soup):
    try:
        exdivamt = soup.find('td', {"data-test": "DIVIDEND_AND_YIELD-value"}).text.strip()
    except:
        exdivamt = "N/A (N/A)"
    return exdivamt


def get_current_price(symbol):
    cprice = 0
    try:
        ticker = yf.Ticker(symbol)
        todays_data = ticker.history(period='1d')
        cprice = todays_data['Close'].iloc[-1]
    except:
        cprice = 0
    return cprice


def get_52weekshigh_price(symbol):
    dataframe = yf.download(symbol, period="1y", auto_adjust=True, prepost=True, threads=True)
    return dataframe['High'].max()


def get_52weekslow_price(symbol):
    dataframe = yf.download(symbol, period="1y", auto_adjust=True, prepost=True, threads=True)
    return dataframe['High'].min()


def get_fromjan2020high_price(symbol):
    dataframe = yf.download(symbol, start='2020-01-01', auto_adjust=True, prepost=True, threads=True)
    return dataframe['High'].max()


def get_fromjan2020low_price(symbol):
    dataframe = yf.download(symbol, period="1y", auto_adjust=True, prepost=True, threads=True)
    return dataframe['High'].min()


def get_market_cap(symbol):
    market_cap_data = int(yf.Ticker(symbol).info['marketCap'])
    #market_cap_data = int(web.get_quote_yahoo(symbol)['marketCap'])
    return market_cap_data


xlfilename = "C:\\Users\\LENOVO\\Desktop\\MY PORTFOLIO.xlsx"
my_wb = openpyxl.load_workbook(xlfilename)
my_sheet_obj = my_wb['NEW WATCHLISTS']
my_row = len([row for row in my_sheet_obj if not all([cell.value == None for cell in row])])
for i in range(2, my_row + 1):
    symbol_cell = my_sheet_obj.cell(row=i, column=1)
    closing_balance = get_current_price(str(symbol_cell.value) + ".NS")
    soup = getSoup(str(symbol_cell.value) + ".NS")
    oneyearhigh_price = get_52weekshigh_price(str(symbol_cell.value) + ".NS")
    oneyearlow_price = get_52weekslow_price(str(symbol_cell.value) + ".NS")
    fromjan2020high_price = get_fromjan2020high_price(str(symbol_cell.value) + ".NS")
    fromjan2020low_price = get_fromjan2020low_price(str(symbol_cell.value) + ".NS")
    market_cap_data = get_market_cap(str(symbol_cell.value) + ".NS")
    ex_dividend_date = get_ex_dividend_date(soup)
    ex_dividend_amt_yield = get_ex_dividend_amt_yield(soup)
    print(ex_dividend_date)
    print(ex_dividend_amt_yield)
    print("{} : {}".format(str(symbol_cell.value), round(closing_balance, 2)))
    my_sheet_obj.cell(row=i, column=7).value = round(closing_balance, 2)
    my_sheet_obj.cell(row=i, column=5).value = round(oneyearhigh_price, 2)
    my_sheet_obj.cell(row=i, column=6).value = round(oneyearlow_price, 2)
    my_sheet_obj.cell(row=i, column=11).value = round(fromjan2020high_price, 2)
    my_sheet_obj.cell(row=i, column=12).value = round(fromjan2020low_price, 2)
    my_sheet_obj.cell(row=i, column=16).value = ex_dividend_date
    my_sheet_obj.cell(row=i, column=17).value = ex_dividend_amt_yield
    # Large-cap companies have a market cap of Rs 20,000 crore or more.
    # the market cap of mid-cap companies is between Rs 5,000 crore and less than Rs 20,000 crore.
    # Small-cap companies have a market cap of below Rs 5,000 crore.
    market_cap = ""
    if market_cap_data < 10000000000:
        market_cap = "MICRO"
    elif market_cap_data > 1000000000 and market_cap_data < 50000000000:
        market_cap = "SMALL"
    elif market_cap_data > 5000000000 and market_cap_data < 200000000000:
        market_cap = "MEDIUM"
    else:
        market_cap = "LARGE"
    my_sheet_obj.cell(row=i, column=3).value = '{:,.2f} ({})'.format(market_cap_data / 10000000, market_cap)
    my_wb.save(xlfilename)
    print("=========SAVED=========")
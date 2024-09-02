import yfinance as yf
import openpyxl
import pandas_datareader as web
import requests
from bs4 import BeautifulSoup
from lxml import etree,html
import datetime


def getSoup(symbol):
    headers = {
        'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36',
        'accept-encoding': None
    }
    url = "https://finance.yahoo.com/quote/" + symbol + "/key-statistics?p=" + symbol + "&.tsrc=fin-srch"
    html = requests.get(url, headers=headers).text
    soup = BeautifulSoup(html, 'html.parser')
    return soup
xlfilename = "C:\\Users\\LENOVO\\OneDrive\\Desktop\\MY PORTFOLIO.xlsx"
my_wb = openpyxl.load_workbook(xlfilename)
my_sheet_obj = my_wb['BONUS AND SPLITS']
my_row = len([row for row in my_sheet_obj if not all([cell.value == None for cell in row])])
for i in range(2, my_row+1):
    exchange_cell = my_sheet_obj.cell(row=i, column=2)
    symbol_cell = my_sheet_obj.cell(row=i, column=3)
    soup = getSoup(str(symbol_cell.value).strip() + ".NS")
    dom = etree.HTML(str(soup))
    #print(html.tostring(dom.xpath('/html/body/div[1]/div/div/div[1]/div/div[3]/div[1]/div/div[1]/div/div/section/div[2]/div[2]/div/div[3]/div/div/table/tbody/*')))
    result = soup.select("#nimbus-app > section > section > section > article > article > div > section:nth-child(2) > div > section:nth-child(3) > table > tbody")
    all_trs = result[0].find_all('tr')
    split_ratio = ''
    split_date = ''
    for tr in all_trs:
        print(tr.find_all('td')[0].text)
        if tr.find_all('td')[0].text.startswith('Last Split Factor'):
            split_ratio = tr.find_all('td')[1].text
        if tr.find_all('td')[0].text.startswith('Last Split Date'):
            split_date = tr.find_all('td')[1].text
    print(symbol_cell.value)
    print(split_ratio)
    print(split_date)
    if split_date != "--" and split_date is not None:
        split_date = datetime.datetime.strptime(split_date, '%m/%d/%Y').strftime("%d-%b-%Y")
    print(split_date)
    my_sheet_obj.cell(row=i, column=5).value = split_ratio
    my_sheet_obj.cell(row=i, column=6).value = split_date
    my_wb.save(xlfilename)
    print("=========SAVED=========")
import datetime
import yfinance as yf
import openpyxl
import xlwings as xw

def get_ticker(symbol):
    return yf.Ticker(symbol)


def format_date(epoch_date):
    return datetime.datetime.utcfromtimestamp(epoch_date).strftime("%d-%b-%Y") if int(epoch_date) != 0 else ''


def amount_to_crores(amount):
    return '{:,.2f}'.format(int(amount) / 10000000) if int(amount) > 0 else 0


def get_price_data(symbol, start_date, price_type='High', stat='max'):
    df = yf.download(symbol, start=start_date, auto_adjust=True, prepost=True, threads=True)
    return df[price_type].max() if stat == 'max' else df[price_type].min()


def update_excel_with_nse_ticker_info(sheet, row, ticker_info):
    def safe_get(info, key, default='0'):
        return info.get(key, default)

    fields = [
        ('address1', 5), ('address2', 6), ('city', 7), ('zip', 8), ('country', 9), ('phone', 10),
        ('fax', 11), ('website', 12), ('industry', 13), ('industryKey', 14), ('industryDisp', 15),
        ('sector', 16), ('sectorKey', 17), ('sectorDisp', 18), ('longBusinessSummary', 19),
        ('fullTimeEmployees', 20), ('auditRisk', 22), ('boardRisk', 23), ('compensationRisk', 24),
        ('shareHolderRightsRisk', 25), ('overallRisk', 26), ('governanceEpochDate', 27, 'date'),
        ('compensationAsOfEpochDate', 28, 'date'), ('maxAge', 29), ('priceHint', 30),
        ('previousClose', 31,'round'), ('open', 32,'round'), ('dayLow', 33,'round'), ('dayHigh', 34,'round'),
        ('regularMarketPreviousClose', 35,'round'), ('regularMarketOpen', 36,'round'), ('regularMarketDayLow', 37,'round'),
        ('regularMarketDayHigh', 38,'round'), ('exDividendDate', 39, 'date'),
        ('fiveYearAvgDividendYield', 40), ('beta', 41), ('trailingPE', 42), ('forwardPE', 43),
        ('volume', 44), ('regularMarketVolume', 45), ('averageVolume', 46),
        ('averageVolume10days', 47), ('averageDailyVolume10Day', 48), ('marketCap', 49),
        ('fiftyTwoWeekLow', 51), ('fiftyTwoWeekHigh', 52), ('priceToSalesTrailing12Months', 53),
        ('fiftyDayAverage', 54), ('twoHundredDayAverage', 55), ('currency', 56),
        ('enterpriseValue', 57, 'crores'), ('profitMargins', 58), ('floatShares', 59),
        ('sharesOutstanding', 60), ('heldPercentInsiders', 61), ('heldPercentInstitutions', 62),
        ('impliedSharesOutstanding', 63, 'crores'), ('bookValue', 64), ('priceToBook', 65),
        ('lastFiscalYearEnd', 66, 'crores'), ('nextFiscalYearEnd', 67, 'crores'),
        ('mostRecentQuarter', 68, 'crores'), ('earningsQuarterlyGrowth', 69, 'crores'),
        ('netIncomeToCommon', 70, 'crores'), ('trailingEps', 71, 'crores'), ('forwardEps', 72),
        ('pegRatio', 73), ('lastSplitFactor', 74), ('lastSplitDate', 75, 'date'),
        ('enterpriseToRevenue', 76, 'crores'), ('enterpriseToEbitda', 77), ('52WeekChange', 78),
        ('SandP52WeekChange', 79), ('lastDividendValue', 80,'round'), ('lastDividendDate', 81, 'date'),
        ('exchange', 82), ('quoteType', 83), ('symbol', 84), ('underlyingSymbol', 85),
        ('shortName', 86), ('longName', 87), ('firstTradeDateEpochUtc', 88, 'date'),
        ('timeZoneFullName', 89), ('timeZoneShortName', 90), ('uuid', 91), ('messageBoardId', 92),
        ('gmtOffSetMilliseconds', 93, 'date'), ('currentPrice', 94,'round'), ('targetHighPrice', 95,'round'),
        ('targetLowPrice', 96,'round'), ('targetMeanPrice', 97), ('targetMedianPrice', 98),
        ('recommendationMean', 99), ('recommendationKey', 100),
        ('numberOfAnalystOpinions', 101), ('totalCash', 102, 'crores'),
        ('totalCashPerShare', 103, 'crores'), ('ebitda', 104, 'crores'), ('totalDebt', 105, 'crores'),
        ('totalRevenue', 106, 'crores'), ('debtToEquity', 107, 'crores'), ('revenuePerShare', 108),
        ('earningsGrowth', 109), ('revenueGrowth', 110), ('grossMargins', 111),
        ('ebitdaMargins', 112), ('operatingMargins', 113), ('financialCurrency', 114),
        ('trailingPegRatio', 115)
    ]

    for field, col, transform in [(f[0], f[1], f[2] if len(f) > 2 else None) for f in fields]:
        value = safe_get(ticker_info, field)
        if transform == 'date':
            value = format_date(value)
        elif transform == 'crores':
            value = amount_to_crores(value)
        elif transform == 'round':
            value = round(float(value),2)
        sheet.cell(row=row, column=col).value = value
    market_cap = ticker_info.get('marketCap',0)
    # Determine market cap category
    if market_cap == 0:
        market_cap_category = ""
    elif market_cap < 10000000000:
        market_cap_category = "MICRO"
    elif market_cap < 50000000000:
        market_cap_category = "SMALL"
    elif market_cap < 200000000000:
        market_cap_category = "MEDIUM"
    else:
        market_cap_category = "LARGE"

    sheet.cell(row=row, column=49).value = amount_to_crores(market_cap)
    sheet.cell(row=row, column=50).value = market_cap_category


def update_excel_with_bse_ticker_info(sheet, row, ticker_info):
    def safe_get(info, key, default='0'):
        return info.get(key, default)

    fields = [
        ('maxAge', 15),
        ('priceHint', 16),
        ('previousClose', 17,'round'),
        ('open', 18,'round'),
        ('dayLow', 19,'round'),
        ('dayHigh', 20,'round'),
        ('regularMarketPreviousClose', 21,'round'),
        ('regularMarketOpen', 22,'round'),
        ('regularMarketDayLow', 23,'round'),
        ('regularMarketDayHigh', 24,'round'),
        ('trailingPE', 25),
        ('forwardPE', 26),
        ('volume', 27),
        ('regularMarketVolume', 28),
        ('averageVolume', 29),
        ('averageVolume10days', 30),
        ('averageDailyVolume10Day', 31),
        ('bid', 32),
        ('ask', 33),
        ('marketCap', 34,'crores'),
        ('fiftyTwoWeekLow', 36),
        ('fiftyTwoWeekHigh', 37),
        ('fiftyDayAverage', 38),
        ('twoHundredDayAverage', 39),
        ('trailingAnnualDividendRate',40),
        ('trailingAnnualDividendYield',41),
        ('currency', 42),
        ('exchange', 43),
        ('quoteType', 44),
        ('symbol', 45),
        ('underlyingSymbol', 46),
        ('shortName', 47),
        ('longName', 48),
        ('firstTradeDateEpochUtc', 49, 'date'),
        ('timeZoneFullName', 50),
        ('timeZoneShortName', 51),
        ('uuid', 52),
        ('gmtOffSetMilliseconds', 53, 'date'),
        ('currentPrice', 54,'round'),
        ('targetHighPrice', 55,'round'),
        ('targetLowPrice', 56,'round'),
        ('targetMeanPrice', 57,'round'),
        ('targetMedianPrice', 58),
        ('recommendationKey', 59),
        ('numberOfAnalystOpinions',60),
        ('grossProfits', 61),
        ('financialCurrency', 62),
        ('trailingPegRatio', 63)
    ]

    for field, col, transform in [(f[0], f[1], f[2] if len(f) > 2 else None) for f in fields]:
        value = safe_get(ticker_info, field)
        if transform == 'date':
            value = format_date(value)
        elif transform == 'crores':
            value = amount_to_crores(value)
        elif transform == 'round':
            value = round(float(value),2)
        sheet.cell(row=row, column=col).value = value

    market_cap = ticker_info.get('marketCap',0)
    # Determine market cap category
    if market_cap == 0:
        market_cap_category = ""
    elif market_cap < 10000000000:
        market_cap_category = "MICRO"
    elif market_cap < 50000000000:
        market_cap_category = "SMALL"
    elif market_cap < 200000000000:
        market_cap_category = "MEDIUM"
    else:
        market_cap_category = "LARGE"

    sheet.cell(row=row, column=34).value = amount_to_crores(market_cap)
    sheet.cell(row=row, column=35).value = market_cap_category


def main():
    xlfilename = "C:\\Users\\LENOVO\\OneDrive\\Desktop\\ALL MARKET ANALYSIS.xlsx"
    try:
        book = xw.Book(xlfilename)
        book.close()
    except Exception as e:
        print(e)
    workbook = openpyxl.load_workbook(xlfilename)
    sheet = workbook['NSE EQUITY']
    total_rows = len([row for row in sheet if not all([cell.value is None for cell in row])])
    print(f"Total Stocks in NSE : {total_rows}")
    '''
    for i in range(2, 3):
        symbol = str(sheet.cell(row=i, column=1).value) + ".NS"
        ticker = get_ticker(symbol)
        ticker_info = ticker.info
        update_excel_with_nse_ticker_info(sheet, i, ticker_info)
        current_price = ticker_info.get('currentPrice', 0)
        print(f"{i - 1}/{total_rows - 1} - {symbol} : {round(float(current_price), 2)}")
        workbook.save(xlfilename)
        print("=========SAVED=========")
    '''
    sheet = workbook['BSE EQUITY']
    total_rows = len([row for row in sheet if not all([cell.value is None for cell in row])])
    print(f"Total Stocks in BSE : {total_rows}")

    for i in range(2, total_rows + 1):
        symbol = str(sheet.cell(row=i, column=3).value) + ".BO"
        ticker = get_ticker(symbol)
        ticker_info = ticker.info
        update_excel_with_bse_ticker_info(sheet, i, ticker_info)
        current_price = ticker_info.get('currentPrice', 0)
        print(f"{i - 1}/{total_rows - 1} - {symbol} : {round(float(current_price), 2)}")
        workbook.save(xlfilename)
        print("=========SAVED=========")

if __name__ == "__main__":
    main()
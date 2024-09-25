import re
import yfinance as yf
import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By

XL_FILE_PATH = "C:\\Users\\LENOVO\\OneDrive\\Desktop\\MY PORTFOLIO.xlsx"
def get_current_nifty():
    return yf.Ticker('^NSEI').history_metadata.get('regularMarketPrice',0)

def get_soup(url):
    headers = {
        'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36',
        'accept-encoding': None
    }
    html = requests.get(url, headers=headers).text
    return BeautifulSoup(html, 'html.parser')


def get_current_nav(soup):
    try:
        # print(soup)
        # print(soup.find_all('div', class_="col l5 fnd2TableDiv fnd2Left"))
        # return float(''.join(filter(lambda x: x.isdigit() or x in '-./\\', soup.find('div', class_="col l5 fnd2TableDiv fnd2Left").find('table').find_all('tr')[0].find('td',class_='fd12Cell contentPrimary bodyLargeHeavy').text.strip())))
        # script = soup.find('script', attrs={'id': '__NEXT_DATA__', 'type': 'application/json'})
        # pattern = re.compile()
        print(soup.select("#fnd2Section > div > div:nth-of-type(1) > div > table > tr:nth-of-type(1) > td.fd12Cell.contentPrimary.bodyLargeHeavy")[0])
        return soup.select_one("#fnd2Section > div > div:nth-of-type(1) > div > table > tr:nth-of-type(1) > td.fd12Cell.contentPrimary.bodyLargeHeavy").text
    except:
        return "N/A"


def get_expense_ratio(soup):
    try:
        return soup.find('h3', class_="ot654subHeading bodyLargeHeavy").text.strip().replace("Expense ratio: ", "")
    except:
        return "N/A"


def process_investments(sheet, start_row, end_row, column_url, column_nav, column_ratio):
    for i in range(start_row, end_row):
        url = str(sheet.cell(row=i, column=column_url).value).strip()
        print(url)
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        driver.get(url)
        current_nav = float(''.join(filter(lambda x: x.isdigit() or x in '-./\\', driver.find_element(By.XPATH,"/html/body/div/div[2]/div[2]/div[1]/div[2]/div/div/div[1]/div/div[3]/div/div/div[1]/div/table/tr[1]/td[2]").text)))
        ## current_nav = get_current_nav(soup)
        expense_ratio = driver.find_element(By.XPATH,"/html/body/div/div[2]/div[2]/div[1]/div[2]/div/div/div[1]/div/div[10]/div/div[3]/h3").text.strip().replace("Expense ratio: ", "")
        current_nifty = get_current_nifty()
        sheet.cell(row=i, column=column_nav).value = current_nav
        sheet.cell(row=i, column=column_ratio).value = expense_ratio
        sheet.cell(row=i, column= 13).value = current_nifty

def process_crypto(sheet, start_row, end_row, column_url, column_price):
    for i in range(start_row, end_row):
        url = str(sheet.cell(row=i, column=column_url).value).strip()
        print(url)
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        driver.get(url)

        html_source = driver.page_source
        soup = BeautifulSoup(html_source, 'html.parser')
        current_price = get_current_crypto_price(soup)
        print(current_price)
        sheet.cell(row=i, column=column_price).value = current_price


def get_current_crypto_price(soup):
    try:
        return float(re.findall("\d+\.\d+",
                                soup.find('div', class_="flex flex-col").find('h1').text.strip().replace(',',
                                                                                                         '').replace(
                                    'Fetch.ai', 'fetchai'))[0])
    except:
        return 0


def main():
    wb = openpyxl.load_workbook(XL_FILE_PATH)
    sheet = wb['ALL INVESTMENTS']

    process_investments(sheet, 19, 24, 7, 4, 2)
    #process_crypto(sheet, 36, 40, 8, 4)

    wb.save(XL_FILE_PATH)
    print("=========SAVED=========")


if __name__ == "__main__":
    main()

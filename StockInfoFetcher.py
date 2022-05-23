import pandas as pd
from bs4 import BeautifulSoup
import re
from selenium import webdriver
import string
import yfinance as yf
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
CHROME_DVR_DIR = 'C:\\YOUTUBE-DL\\chromedriver.exe'
#pd.options.display.float_format = '{:,.2f}'.format
xlfilename = "C:\\Users\\Sambit Samal\\Desktop\\PORTPOLIO ANALYSIS.xlsx"
def yahoo_financial_statements(ticker):
    is_link = f'https://finance.yahoo.com/quote/{ticker}/financials?p={ticker}'
    bs_link = f'https://finance.yahoo.com/quote/{ticker}/balance-sheet?p={ticker}'
    cf_link = f'https://finance.yahoo.com/quote/{ticker}/cash-flow?p={ticker}'
    statements_list = [is_link,bs_link,cf_link]
    headers = []
    temp_list = []
    label_list = []
    final = []
    index = 0
    df_lists = list()
    driver = webdriver.Chrome(CHROME_DVR_DIR)
    for link in statements_list:
        driver.get(link)
        html = driver.execute_script('return document.body.innerHTML;')
        soup = BeautifulSoup(html,'lxml')
        features = soup.find_all('div', class_='D(tbr)')
        #create headers
        for item in features[0].find_all('div', class_='D(ib)'):
            headers.append(item.text)
        #statement contents
        while index <= len(features)-1:
            #filter for each line of the statement
            temp = features[index].find_all('div', class_='D(tbc)')
            for line in temp:
                #each item added to a temp list
                temp_list.append(line.text)
            #temp_list added to final list
            final.append(temp_list)
            #clear temp_list
            temp_list = []
            index+=1
        df = pd.DataFrame(final[1:])
        df.columns = headers
        #df.index = final[1]
        #display(df.iloc[:, [0,1,5,4,3,2]])
        #function to make all values numerical
        def convert_to_numeric(column):
            first_col = [i.replace(',','') for i in column]
            second_col = [i.replace('-','') for i in first_col]
            final_col = ['{:,.2f}'.format((0 if x == '' else float(x))/10000000) for x in second_col]
            return final_col
        for column in headers[1:]:
            df[column] = convert_to_numeric(df[column])
        final_df = df.fillna('-')
        df_lists.append(final_df)
        #reset all lists
        headers = []
        temp_list = []
        label_list = []
        final = []
        index = 0
    driver.close()
    return df_lists

def multiple_dfs(my_wb,df_list,header_list, sheets, file_name, spaces):
    if not sheets in my_wb.sheetnames:
        my_wb.create_sheet(sheets)
    my_sheet_obj = my_wb[sheets]
    num_row = 1
    index = 0
    for dataframe in df_list:
        header= header_list[index]
        dataframe.head().style.format("{:,.2f}")
        dataframe.rename(columns={'Breakdown': header+' - Breakdown'}, inplace=True)
        rows = dataframe_to_rows(dataframe)
        for r_idx, row in enumerate(rows, num_row):
            for c_idx, val in enumerate(row, 1):
                my_sheet_obj.cell(row=r_idx, column=c_idx).value=str(val)
        num_row = num_row + dataframe.shape[0] + spaces + 1
        index = index + 1
    my_wb.save(xlfilename)


#financials = yahoo_financial_statements()
#income_statement = financials[0]
#balance_sheet = financials[1]
#cash_flow_statement = financials[2]
#income_statement.index = income_statement[['Breakdown']]
def writeStockInfo():
    my_wb = openpyxl.load_workbook(xlfilename)
    my_sheet_obj = my_wb['ALL STOCKS']
    my_row = len([row for row in my_sheet_obj if not all([cell.value == None for cell in row])])
    stocks_info_list = []
    for i in range(2, my_row+1):
        exchange_cell = my_sheet_obj.cell(row = i, column = 2)
        symbol_cell = my_sheet_obj.cell(row = i, column = 3)
        ticker = str(symbol_cell.value)+"."+("NS" if exchange_cell.value == "NSE" else "BO")
        print(str(symbol_cell.value))
        ticker_info = yf.Ticker(ticker)
        #json_info = json.dumps(ticker_info.info)
        ticker_data = {key.upper():value for key, value in ticker_info.info.items()}
        stocks_info_list.append(ticker_data)
        #stock_info = pd.json_normalize(json_data)
    #stock_info = stock_info.T
    #stock_info.T.columns = ['Stock Points','Value']
    #print(stock_info.T)
    json_info = json.dumps(stocks_info_list).replace(", nan," , ', "null",')
    json_data = json.loads(json_info)
    for element in json_data:
        element.pop('ADDRESS1', None)
        element.pop('ADDRESS2', None)
        element.pop('COMPANYOFFICERS', None)
    stock_info = pd.json_normalize(json_data)
    print(stock_info)
    rows = dataframe_to_rows(stock_info,index = False)
    for i in range(1, my_row+1):
        for r_idx, row in enumerate(rows, i):
            for c_idx, val in enumerate(row, 5):
                my_sheet_obj.cell(row=r_idx, column=c_idx).value=str(val)
    my_wb.save(xlfilename)
def writeStockStatement():
    my_wb = openpyxl.load_workbook(xlfilename)
    my_sheet_obj = my_wb['ALL STOCKS']
    my_row = len([row for row in my_sheet_obj if not all([cell.value == None for cell in row])])
    stocks_info_list = []
    for i in range(2, my_row+1):
        exchange_cell = my_sheet_obj.cell(row = i, column = 2)
        symbol_cell = my_sheet_obj.cell(row = i, column = 3)
        ticker = str(symbol_cell.value)+"."+("NS" if exchange_cell.value == "NSE" else "BO")
        print(ticker)
        financials = yahoo_financial_statements(ticker)
        income_statement = financials[0]
        balance_sheet = financials[1]
        cash_flow_statement = financials[2]
        print('income_statement')
        print(income_statement)
        print('balance_sheet')
        print(balance_sheet)
        print('cash_flow_statement')
        print(cash_flow_statement)
        #income_statement.index = income_statement[['Breakdown']]
        dfs = [income_statement,balance_sheet,cash_flow_statement]
        header_list = ["Income Statement","Balance Sheet","Cash Flow Statement"]
        multiple_dfs(my_wb,dfs,header_list, ticker, xlfilename, 4)
        
from datetime import datetime
today = datetime.now()
#writeStockInfo()
writeStockStatement()
#if today.day == 1:
    #writeStockInfo()
#    writeStockStatement()
#else:
#    writeStockInfo()

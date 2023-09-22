import re
import requests
import os
import io
import glob
import fileinput
import openpyxl
import urllib3
from openpyxl.drawing.image import Image
import PIL
from bs4 import BeautifulSoup
from requests.exceptions import ConnectionError
proxy = {'https': 'http://127.0.0.1:9666'} 
def appendToXls(my_wb,xlfilename,my_sheet_obj,linklist):
    current_first_col = []
    my_row = my_sheet_obj.max_row
    for i in range(1, my_row + 1):
        cell_obj = my_sheet_obj.cell(row = i, column = 1)
        current_first_col.append(cell_obj.value)
    for i in range(1, my_row + 1):
        for link in linklist:
            if link not in current_first_col:
                first_col=link
                print("{} ; Completed : {}/{}".format(link,linklist.index(link)+1,len(linklist)))
                try:
                    r = requests.get(link,proxies = proxy)
                except ConnectionError as e:
                    my_sheet_obj.append([link,'=HYPERLINK("{}", "{}")'.format(link, link),"--------------------"])
                    my_wb.save(xlfilename)
                    continue
                if r.content.lower() == b'invalid':
                    my_sheet_obj.append([link,'=HYPERLINK("{}", "{}")'.format(link, link),"--------------------"])
                    my_wb.save(xlfilename)
                    continue
                if r.status_code != 200:
                    my_sheet_obj.append([link,'=HYPERLINK("{}", "{}")'.format(link, link),"--------------------"])
                    my_wb.save(xlfilename)
                    continue
                second_col=r.url
                soup = BeautifulSoup(r.content, 'html.parser')
                third_col=soup.select_one('title').text
                my_sheet_obj.append([first_col,'=HYPERLINK("{}", "{}")'.format(second_col, second_col),third_col])        
                my_wb.save(xlfilename)
for filename in glob.glob(os.path.join('E:\\telegram-data\\', '*.json')):
    with fileinput.FileInput(filename, inplace=True) as file:
        for line in file:
            print(line.replace("\\n", " "), end='')
    print("Working on "+filename)
    import re
    linklist = []
    with open(filename) as fh: 
        for line in fh: 
            match = re.findall("(?s)(?<=https).+?(?= )", line) 
            for m in match:
                linklist.append(("https"+m).replace("\",",""))
    seen = set()
    linklist =  [x for x in linklist if x not in seen and not seen.add(x)]            
    xlfilename = "E:\\telegram-data\\ALLTELEGRAMLINKS.xlsx"
    if not os.path.isfile(xlfilename):
        my_wb = openpyxl.Workbook()
        my_sheet_obj = my_wb.active
        appendToXls(my_wb,xlfilename,my_sheet_obj,linklist)
    else:
        my_wb = openpyxl.load_workbook(xlfilename)
        my_sheet_obj = my_wb.active
        appendToXls(my_wb,xlfilename,my_sheet_obj,linklist)
        
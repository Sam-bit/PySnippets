import yfinance as yf
import openpyxl
import pandas_datareader as web

def get_current_price(symbol):
    ticker = yf.Ticker(symbol)
    todays_data = ticker.history(period='1d')
    return todays_data['Close'][0]

def get_market_cap(symbol):
    market_cap_data = int(web.get_quote_yahoo(symbol)['marketCap'])
    return market_cap_data
xlfilename = "C:\\Users\\Sambit Samal\\Desktop\\MY PORTFOLIO.xlsx"
my_wb = openpyxl.load_workbook(xlfilename)
my_sheet_obj = my_wb['CURRENT PORTFOLIO']
my_row = len([row for row in my_sheet_obj if not all([cell.value == None for cell in row])])
for i in range(2, my_row+1):
    exchange_cell = my_sheet_obj.cell(row = i, column = 2)
    symbol_cell = my_sheet_obj.cell(row = i, column = 3)
    closing_balance=get_current_price(str(symbol_cell.value)+"."+("NS" if exchange_cell.value == "NSE" else "BO"))
    market_cap_data = get_market_cap(str(symbol_cell.value)+"."+("NS" if exchange_cell.value == "NSE" else "BO"))
    print("{} : {}".format(str(symbol_cell.value), round(closing_balance, 2)))
    my_sheet_obj.cell(row=i, column=9).value = round(closing_balance, 2)
    #Large-cap companies have a market cap of Rs 20,000 crore or more. 
    #the market cap of mid-cap companies is between Rs 5,000 crore and less than Rs 20,000 crore. 
    #Small-cap companies have a market cap of below Rs 5,000 crore.
    market_cap = ""
    if market_cap_data < 50000000000:
        market_cap = "SMALL"
    elif market_cap_data > 5000000000 and market_cap_data < 200000000000:
        market_cap = "MEDIUM"
    else:
        market_cap = "LARGE"
    my_sheet_obj.cell(row=i, column=6).value = '{:,.2f} ({})'.format(market_cap_data/10000000,market_cap)
    my_wb.save(xlfilename)